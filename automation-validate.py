from selenium import webdriver
import time
from openpyxl import load_workbook

def get_ids(filename):

	ids = []

	wb = load_workbook(filename = filename)
	sheet = wb.get_active_sheet()

	col = 1
	for row in range(1,sheet.max_row+1):
		ids.append(int(sheet.cell(column = col, row = row).value))

	return ids

if __name__ == "__main__":

	# display = Display(visible=0, size=(1366, 768))
	browser = webdriver.Firefox()
	url = 'https://testnetbankadm-app4.moon.profis/eloyal/html/eng/netbossadm.html'
	browser.get(url)
	time.sleep(30)

	c_maintenance = browser.find_element_by_id('serv_link1')
	c_maintenance.click()

	ids = get_ids('test.xlsx')

	for id_ in ids:
		iframe = browser.find_element_by_tag_name("iframe")
		browser.switch_to_frame(iframe)
		time.sleep(2)

		c_number = browser.find_element_by_xpath('//input[@name="ownum"]')
		c_number.send_keys(id_)
		time.sleep(2)

		c_number_s_btn = browser.find_element_by_name('searchownum')
		c_number_s_btn.click()
		time.sleep(3)

		vldt_btn = browser.find_element_by_id('validcont')
		vldt_btn.click()
		time.sleep(3)

		browser.switch_to_default_content()

		c_maintenance = browser.find_element_by_id('serv_link1')
		c_maintenance.click()
		time.sleep(3)
	
	browser.close()

