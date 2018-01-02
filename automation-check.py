from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from selenium.webdriver.support.ui import Select
import time

def get_ids(filename):

	ids = []

	wb = load_workbook(filename = filename)
	sheet = wb.get_active_sheet()

	col = 1
	for row in range(1,sheet.max_row+1):
		ids.append(int(sheet.cell(column = col, row = row).value))

	return ids

def change_signature(function):
	function.select_by_index(0)

if __name__ == "__main__":

	# display = Display(visible=0, size=(1366, 768))
	browser = webdriver.Firefox()
	url = 'https://testnetbankadm-app4.moon.profis/eloyal/html/eng/netbossadm.html'
	browser.get(url)
	time.sleep(30)
	ids = get_ids('test.xlsx')

	c_maintenance = browser.find_element_by_id('serv_link1')
	c_maintenance.click()
	time.sleep(3)

	for id_ in ids:

		iframe = browser.find_element_by_tag_name("iframe")
		browser.switch_to_default_content()
		browser.switch_to_frame(iframe)
		time.sleep(3)

		c_number = browser.find_element_by_xpath('//input[@name="ownum"]')
		c_number.send_keys(id_)

		c_number_s_btn = browser.find_element_by_name('searchownum')

		c_number_s_btn.click()

		time.sleep(3)

		back_to_wip = browser.find_element_by_name('wipstate')
		back_to_wip.click()

		deposit_func = Select(browser.find_element_by_xpath("//select[@name='sfuncval1']"))
		internal_trns = Select(browser.find_element_by_xpath("//select[@name='sfuncval3']"))
		external_trns = Select(browser.find_element_by_xpath("//select[@name='sfuncval4']"))
		standing_ordrs = Select(browser.find_element_by_xpath("//select[@name='sfuncval6']"))
		digipass_func = Select(browser.find_element_by_xpath("//select[@name='sfuncval7']"))
		cps_func = Select(browser.find_element_by_xpath("//select[@name='sfuncval8']"))
		invest_func = Select(browser.find_element_by_xpath("//select[@name='sfuncval9']"))

		functions = [deposit_func,internal_trns,external_trns,standing_ordrs,digipass_func,cps_func,invest_func]

		for f in functions:
			change_signature(f)
		
		try: 
			check_btn = browser.find_element_by_id('checkcont')
			check_btn.click()
			time.sleep(3)
		except Exception as e:
			print e
			pass

		browser.switch_to_default_content()

		c_maintenance = browser.find_element_by_id('serv_link1')
		c_maintenance.click()
		time.sleep(3)

	browser.close()

