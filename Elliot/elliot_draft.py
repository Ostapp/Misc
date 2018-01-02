import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from user_agent import generate_user_agent
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
import json
import random

# sudo docker run -p 5023:5023 -p 8050:8050 -p 8051:8051 scrapinghub/splash --max-timeout 120


def get_cities():

	cities = []

	wb = load_workbook(filename = 'France 50 biggest cities.xlsx')
	sheet = wb.get_active_sheet()

	col = 1
	for row in range(1,55):
		cities.append(sheet.cell(column = col, row = row).value)

	return cities

def open_proxies_list():
	with open('proxy-list.txt','r') as f:
		proxies = f.readlines()

	stripped_proxies = []

	for proxy in proxies:
		proxy = proxy.strip()
		stripped_proxies.append(proxy)

	return stripped_proxies

def headers_ua():

	headers = requests.utils.default_headers()

	headers.update({
		'User-Agent': random_user_agent()
		})
	return headers



cities = get_cities()

city = cities[0]

proxy_url = "http://pubproxy.com/api/proxy?format=json&user_agent=true"
proxy_response = requests.get(proxy_url)
proxy_dict = json.loads(proxy_response.content)
proxy = {'http':str(proxy_dict['data'][0]['ipPort'])}


url = 'https://www.pagesjaunes.fr/annuaire/chercherlespros?quoiqui=pharmacie&ou='+city+'&proximite=0&quoiQuiInterprete=pharmacie'
splash = 'http://localhost:8050/render.html'
# user_agent = generate_user_agent()
# headers = {'User-Agent': user_agent}
headers = headers_ua()
# url_h = 'http://xhaus.com/headers'
page = requests.get(url,headers=headers)

browser = webdriver.Chrome()
browser.get(url)
details= browser.find_elements_by_class_name('details-links ')

def center_on_button(element):

	location_x = element.location['x']
	location_y = element.location['y']
	
	browser.execute_script("window.scrollTo(0, %d);" %location_x)
	browser.execute_script("window.scrollTo(0, %d);" %location_y)

	i = ActionChains(browser).move_to_element(element)
	i.perform()

def click(browser, element):

	ActionChains(browser).click(element).perform()

link_to_website = str(browser.find_element_by_class_name('bloc-info-sites-reseaux').find_element_by_class_name('value').text)



soup = BeautifulSoup(page.text)

results = soup(class_='details-links ')[1:]

more_info = results[0].h2.a

cities = get_cities()

city = cities[0]

proxy_url = "http://pubproxy.com/api/proxy?format=json&user_agent=true"
proxy_response = requests.get(proxy_url)
proxy_dict = json.loads(proxy_response.content)
proxy = {'http':str(proxy_dict['data'][0]['ipPort'])}


url = 'https://www.pagesjaunes.fr/annuaire/chercherlespros?quoiqui=pharmacie&ou='+city+'&proximite=0&quoiQuiInterprete=pharmacie'
splash = 'http://localhost:8050/render.html'
# user_agent = generate_user_agent()
# headers = {'User-Agent': user_agent}
headers = headers_ua()
# url_h = 'http://xhaus.com/headers'
page = requests.get(url,headers=headers)

chrome_options = webdriver.ChromeOptions()
prefs = {"profile.default_content_setting_values.notifications" : 2}
chrome_options.add_experimental_option("prefs",prefs)
browser = webdriver.Chrome(chrome_options=chrome_options)
browser.get(url)
browser.find_element_by_xpath('/html/body/div[1]/div/button').click()

soup = BeautifulSoup(page.text)

results = soup(class_='details-links ')[1:]

more_info = results[0].h2.a
