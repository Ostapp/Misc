import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from user_agent import generate_user_agent
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import json
import random
import time
import subprocess
import pickle
import re
import xlsxwriter
# sudo docker run -p 5023:5023 -p 8050:8050 -p 8051:8051 scrapinghub/splash --max-timeout 120

class Administrator():

	def __init__(self):
		self.cities = self.get_cities()
		# self.proxies = self.open_proxies_list()
		# self.proxy_gen = self.get_proxy()
		self.headers = self.headers_ua()
		self.start_urls = self.start_search_urls()
		self.links = {}
		self.spider = Spider()

	def get_cities(self):

		cities = []

		wb = load_workbook(filename = 'France 50 biggest cities.xlsx')
		sheet = wb.get_active_sheet()

		col = 1
		for row in range(1,55):
			cities.append(sheet.cell(column = col, row = row).value)

		return cities

	def open_proxies_list(self):
		with open('thebigproxylist-17-09-03.txt','r') as f:
			proxies = f.readlines()

		stripped_proxies = []

		for proxy in proxies:
			proxy = proxy.strip()
			stripped_proxies.append(proxy)

		return stripped_proxies

	# def get_proxy(self):
	# 	proxies = self.open_proxies_list()
	# 	for proxy in proxies[1:]

	def random_user_agent(self):
		with open('user-agents.txt','r') as f:
			user_agents = f.readlines()
		user_agents = [h.rstrip('\n') for h in user_agents]
		random_index = random.randint(0,len(user_agents)-1)
		ua = user_agents[random_index]
		return ua

	def headers_ua(self):

		headers = requests.utils.default_headers()

		headers.update({
			'User-Agent': self.random_user_agent()
			})
		return headers

	def start_search_urls(self):

		urls = []

		for city in self.cities:
			url = 'https://www.pagesjaunes.fr/annuaire/chercherlespros?quoiqui=pharmacie&ou='+city+'&proximite=0&quoiQuiInterprete=pharmacie'

			urls.append(url)

		return urls

	def process_start_url(self,url):

		print "processing " + url
		master_page = SearchPage(url)
		num_of_entries = len(master_page.names)
		temp_saved_records = {}
		temp_already_saved = 0

		while len(temp_saved_records) < num_of_entries:
			work_page = SearchPage(url)	
			for index in range(num_of_entries):
				random.shuffle(work_page.details_links)
				spider.center_on_button(work_page.browser, work_page.details_links[index])
				time.sleep(random.randint(1,5))

				name = work_page.names[index].text

				if name not in self.links:
					print name + " NAME IS UNIQUE"

					while work_page.browser.current_url == master_page.browser.current_url:
						try:
							spider.click(work_page.browser, work_page.details_links[index])
						except Exception as e:
							work_page.close_pop_ups(work_page.browser)
							try:
								spider.click(work_page.browser, work_page.details_links[index])
							except:
								pass
							pass

					time.sleep(random.randint(1,5))

					info_page = InfoPage(work_page.browser)
					try:
						self.links[name] = info_page.link_to_website
						temp_saved_records[name] = info_page.link_to_website
						print str(len(temp_saved_records))
						print "LINK "+info_page.link_to_website+" APPENDED"
					except Exception as e:
						print e
						pass
					info_page.browser.close()
					time.sleep(1)
					work_page = SearchPage(url)

				elif temp_already_saved >= 20:
					print "THE PAGE HAS BEEN EXHAUSTED"
					work_page.browser.close()
					break
				else:
					print "*"*10 + "ENTRY ALREADY EXISTS" + "*"*10
					temp_already_saved+=1
					continue
			else:
				break
			try:
				work_page.browser.close()
			except Exception as e:
				print e
				pass

		if master_page.next_page:
			spider.center_on_button(master_page.browser,master_page.next_page)
			prev_url = master_page.browser.current_url
			while prev_url == master_page.browser.current_url:
				try:
					master_page.close_pop_ups(master_page.browser)
					spider.click(master_page.browser, master_page.next_page)
				except Exception as e:
					print e
					spider.click(master_page.browser, master_page.next_page)
			time.sleep(5)
			url = re.sub('contexte.*&','',master_page.browser.current_url)
			print "next page url " + url
			master_page.browser.close()
			try:
				work_page.browser.close()
			except:
				pass
			self.process_start_url(url)
		else:
			"*"*10 + "no next page" + "*"*10
			master_page.browser.close()
			admin.save()
			try:
				work_page.browser.close()
			except:
				pass
			return

	def start(self):
		for url in self.start_urls[-24::-1]:
			self.process_start_url(url)

	def save(self):
		with open ('results', 'w') as f:
			pickle.dump(admin.links, f)

	def load(self):
		with open ('results','r') as f:
			results = pickle.load(f)
			return results
	
	def make_xlsx_file(self):

		workbook = xlsxwriter.Workbook('Elliot.xlsx')
		worksheet = workbook.add_worksheet()

		results = {value for value in self.links.values()}

		row = 0
		for result in results:
			worksheet.write_string(row, 0, result)
			row +=1
		workbook.close()

	def save_unique_results(self):
		results_to_process = self.load()
		results = {value for value in results_to_process.values()}


class Page(object):

	def __init__(self, url=None, render=True):
		if render == True:
			self.browser = self.render_page(url)
		self.url = url

	def close_pop_ups(self, browser):
		try:
			browser.find_element_by_xpath('/html/body/div[1]/div/button/span').click()
		except:
			pass
		try:
			browser.find_element_by_xpath('/html/body/div[1]/div/button').click()
		except:
			pass

		try:
			browser.find_element_by_id('acc-alert-close').click()
		except:
			pass
		try:
			close = browser.find_elements_by_id('kamClose')
			for button in close:
				button.click()
		except:
			pass
		try:
			browser.find_element_by_class_name('pjpopin-closer-grandePopin').click()
		except:
			pass
		try:
			browser.find_element_by_class_name('pjpopin-closer').click()
		except:
			pass
		try:
			browser.find_element_by_class_name('kclose').click()
		except:
			pass
		try: 
			browser.find_element_by_class_name('lien-fermer').click()
		except:
			pass

	def render_page(self, url):

		def run_proxy_browser(url):

			def restart_vpn():

				try:
					browser.close()
				except:
					pass
				try:
					os.system('''ps axf | grep hma-vpn.sh | grep -v grep | awk '{print "kill " $1 }' | sh''')
					time.sleep(10)
				except: 
					pass
				try:
					subprocess.Popen(["exec gnome-terminal -e 'bash -c \"sudo bash hma-vpn.sh -c id-file; exec bash\"'"], stdout=subprocess.PIPE, shell=True)
					time.sleep(10)
				except:
					return restart_vpn()

			# PROXY = next(admin.proxy_gen)
			# print PROXY
			profile = webdriver.FirefoxProfile()
			profile.set_preference("general.useragent.override",admin.random_user_agent())
			profile.set_preference("http.response.timeout", 30)
			profile.set_preference("dom.max_script_run_time", 30)
			webdriver.DesiredCapabilities.FIREFOX['marionette'] = False
			# webdriver.DesiredCapabilities.FIREFOX['proxy']={
			# 	"httpProxy":PROXY,
			# 	"ftpProxy":PROXY,
			# 	"sslProxy":PROXY,
			# 	"noProxy":None,
			# 	"proxyType":"MANUAL",
			# 	"autodetect":False
			# }

			browser = webdriver.Firefox(profile)
			browser.set_page_load_timeout(120)

			try:
				browser.get(url)
			except Exception as e:
				print e
				browser.close()
				return run_proxy_browser(url)

			if browser.title == u'Problem loading page':
				print "Problem loading page"
				browser.close()
				return run_proxy_browser(url)
			elif browser.title == u'You have been blocked':
				print u'You have been blocked'
				restart_vpn()
				return run_proxy_browser(url)
			elif u"JE SUIS UN HUMAIN" in browser.page_source:
				print "JE SUIS UN HUMAIN"
				restart_vpn()
				return run_proxy_browser(url)
			elif u'PAGESJAUNES.FR protège son contenu contre les robots et le réserve aux êtres humains.' in browser.page_source:
				print "JE SUIS UN HUMAIN"
				restart_vpn()
				return run_proxy_browser(url)
			try:
				if len(browser.page_source) < 100:
					browser.close()
					restart_vpn()
					return run_proxy_browser(url)
				else:
					pass
			except:
				return run_proxy_browser(url)

			self.close_pop_ups(browser)
			browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
			time.sleep(3)

			return browser

		time.sleep(random.randint(1,5))
		browser = run_proxy_browser(url)

		return browser

class SearchPage(Page):

	def __init__(self, url):
		super(SearchPage, self).__init__(url,render=True)
		try:
			self.details_links = self.browser.find_elements_by_class_name('details-links ')
			self.names = self.browser.find_elements_by_class_name('denomination-links')
		except Exception as e:
			print e
			self.browser.close()
			self.browser = self.render_page(url)
		try:
			self.next_page = self.browser.find_element_by_id('pagination-next')
		except Exception as e:
			print e
			print "*"*10 + "no next page during rendering page detected" + "*"*10
			self.next_page = 0
		
class InfoPage(Page):
	
	def __init__(self, browser):
		super(InfoPage, self).__init__(render=False)
		self.browser = browser
		self.browser.set_page_load_timeout(60)
		self.height = random.randint(50,400)
		self.browser.execute_script("window.scrollTo(0, "+ str(self.height) + ");")
		try:
			self.link_to_website = str(self.browser.find_element_by_class_name('bloc-info-sites-reseaux').find_element_by_class_name('value').text)
		except:
			self.link_to_website = "no link to website available"

class Spider():

	def center_on_button(self, browser, element):

		location_x = element.location['x']
		location_y = element.location['y'] - 200
		
		browser.execute_script("window.scrollTo(0, %d);" %location_x)
		browser.execute_script("window.scrollTo(0, %d);" %location_y)

	def click(self, browser, element):
		ActionChains(browser).click(element).perform()

	def center_on_button_and_click(self, browser, element):
		ActionChains(browser).move_to_element(element).click().perform()


if __name__ == "__main__":
	# subprocess.Popen(["exec gnome-terminal -e 'bash -c \"sudo bash hma-vpn.sh -c id-file; exec bash\"'"], stdout=subprocess.PIPE, shell=True)

	spider = Spider()
	admin = Administrator()
	admin.links = admin.load()
	admin.start()

#nantes



